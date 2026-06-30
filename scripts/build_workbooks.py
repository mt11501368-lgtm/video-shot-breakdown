#!/usr/bin/env python3
import argparse
import json
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


NAVY = "172033"
PALE = "EEF4FF"
BLUE = "2D6CDF"
GRID = Side(style="thin", color="D7DDE8")


def value(data, key, default=""):
    result = data.get(key, default)
    if result is None:
        return ""
    result = str(result)
    return "'" + result if result.startswith(("=", "+", "-", "@")) else result


def parse_time(text):
    parts = [float(x) for x in re.findall(r"\d+(?:\.\d+)?", str(text))]
    if len(parts) == 1:
        return parts[0]
    if len(parts) == 2:
        return parts[0] * 60 + parts[1]
    return parts[-3] * 3600 + parts[-2] * 60 + parts[-1]


def duration(shot):
    try:
        return f"{max(0, parse_time(shot['end']) - parse_time(shot['start'])):.1f} 秒"
    except (KeyError, ValueError, TypeError):
        return value(shot, "duration")


def resolve_image(raw, base):
    if not raw:
        return None
    path = Path(raw).expanduser()
    if not path.is_absolute():
        path = base / path
    return path.resolve() if path.exists() else None


def add_title(ws, title, subtitle, columns):
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=columns)
    cell = ws.cell(1, 1, title)
    cell.font = Font(size=20, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=columns)
    sub = ws.cell(3, 1, subtitle)
    sub.font = Font(size=11, italic=True, color="40506A")
    sub.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 32


def write_table(ws, headers, rows, widths, start_row=1, image_col=None, row_height=64):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(start_row, col, header)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=GRID, right=GRID, top=GRID, bottom=GRID)
    ws.row_dimensions[start_row].height = 35

    embedded = 0
    for row_no, row in enumerate(rows, start_row + 1):
        for col_no, item in enumerate(row, 1):
            cell = ws.cell(row_no, col_no, "" if isinstance(item, Path) else item)
            cell.alignment = Alignment(horizontal="center" if col_no <= 4 else "left", vertical="top", wrap_text=True)
            cell.border = Border(left=GRID, right=GRID, top=GRID, bottom=GRID)
            if row_no % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=PALE)
            if col_no == 1:
                cell.font = Font(bold=True, color=BLUE)
        if image_col and isinstance(row[image_col - 1], Path):
            image = Image(str(row[image_col - 1]))
            image.width, image.height = 220, 102
            ws.add_image(image, f"{get_column_letter(image_col)}{row_no}")
            ws.row_dimensions[row_no].height = 79
            embedded += 1
        else:
            ws.row_dimensions[row_no].height = row_height

    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = f"A{start_row + 1}"
    if rows:
        ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{start_row + len(rows)}"
    ws.sheet_view.showGridLines = False
    return embedded


def finish(wb, path, title):
    for ws in wb.worksheets:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.oddFooter.center.text = "第 &P 页 / 共 &N 页"
    wb.properties.title = title
    wb.properties.creator = "Codex"
    wb.save(path)


def cover_rows(video, shot_count, purpose):
    return [
        ["视频", value(video, "title"), "来源", value(video, "source"), "镜头数", str(shot_count)],
        ["时长", value(video, "duration"), "画面", value(video, "resolution"), "帧率", value(video, "fps")],
        ["类型", value(video, "type"), "用途", purpose, "输出", "带关键帧的 Excel 拉片表"],
    ]


def create_learning(payload, json_dir, output):
    shots = payload.get("shots", [])
    wb = Workbook()
    cover = wb.active
    cover.title = "学习拆解说明"
    add_title(cover, "学习拆解版拉片表", "核心问题：它为什么有效？分析叙事、镜头、节奏、声音、情绪与可迁移规律。", 6)
    write_table(cover, ["项目", "内容", "项目", "内容", "项目", "内容"], cover_rows(payload.get("video", {}), len(shots), "学习规律并迁移到新题材"), [15, 38, 15, 38, 15, 38], start_row=5, row_height=44)

    main = wb.create_sheet("学习拉片表")
    rows = []
    expected_images = 0
    for shot in shots:
        image = resolve_image(shot.get("keyframe"), json_dir)
        expected_images += int(image is not None)
        rows.append([
            value(shot, "no"), f"{value(shot, 'start')}–{value(shot, 'end')}", image or value(shot, "keyframe"),
            value(shot, "narrative_role"), value(shot, "visual"), value(shot, "shot_size"), value(shot, "camera"),
            value(shot, "sound"), value(shot, "emotion"), value(shot, "why_effective"), value(shot, "transferable_formula"),
        ])
    embedded = write_table(main, ["镜号", "时间段", "关键帧", "叙事功能", "画面与构图", "景别", "运镜/剪辑", "音乐/音效", "情绪作用", "为什么有效", "可迁移方法"], rows, [8, 17, 33, 25, 42, 15, 20, 42, 24, 44, 48], image_col=3)

    structure = wb.create_sheet("结构总结")
    srows = [[value(x, "shot_range"), value(x, "time_range"), value(x, "task"), value(x, "emotion"), value(x, "sound"), value(x, "formula")] for x in payload.get("structure", [])]
    write_table(structure, ["镜头范围", "时间段", "结构任务", "情绪曲线", "声音策略", "可复用公式"], srows, [15, 22, 28, 31, 48, 54], row_height=72)

    techniques = wb.create_sheet("技巧清单")
    trows = [[value(x, "category"), value(x, "name"), value(x, "evidence"), value(x, "formula")] for x in payload.get("techniques", [])]
    write_table(techniques, ["技巧类型", "技巧名称", "原片证据", "迁移公式"], trows, [18, 28, 58, 58], row_height=62)
    finish(wb, output, f"{value(payload.get('video', {}), 'title', '视频')}—学习拆解版")
    return expected_images, embedded


def create_replica(payload, json_dir, output):
    shots = payload.get("shots", [])
    wb = Workbook()
    cover = wb.active
    cover.title = "复刻执行说明"
    add_title(cover, "模仿复刻版拉片表", "核心问题：我怎么拍出类似效果？直接用于拍摄、AI生成、配音与剪辑执行。", 6)
    write_table(cover, ["项目", "内容", "项目", "内容", "项目", "内容"], cover_rows(payload.get("video", {}), len(shots), "执行拍摄与剪辑"), [15, 38, 15, 38, 15, 38], start_row=5, row_height=44)

    storyboard = wb.create_sheet("分镜脚本")
    rows = []
    expected_images = 0
    for shot in shots:
        image = resolve_image(shot.get("keyframe"), json_dir)
        expected_images += int(image is not None)
        rows.append([
            value(shot, "no"), f"{value(shot, 'start')}–{value(shot, 'end')}", duration(shot), image or value(shot, "keyframe"),
            value(shot, "visual"), value(shot, "action"), value(shot, "shot_size"), value(shot, "camera"),
            value(shot, "edit"), value(shot, "dialogue"), value(shot, "sound"), value(shot, "ai_prompt"),
        ])
    embedded = write_table(storyboard, ["镜号", "时间段", "拍多久", "参考帧", "拍什么", "演员动作/调度", "景别/机位", "怎么运镜", "怎么剪", "台词/旁白", "音乐音效", "AI生成提示词"], rows, [8, 17, 11, 33, 41, 45, 18, 20, 42, 38, 45, 58], image_col=4)

    checklist = wb.create_sheet("拍摄清单")
    crows = [[value(x, "category"), value(x, "item"), value(x, "details"), value(x, "shots")] for x in payload.get("checklist", [])]
    write_table(checklist, ["类别", "项目", "准备内容", "适用镜头"], crows, [16, 28, 78, 20], row_height=58)

    dialogue = wb.create_sheet("台词文案")
    drows = [[value(x, "time_range"), value(x, "shot"), value(x, "speaker"), value(x, "line"), value(x, "delivery"), value(x, "subtitle")] for x in payload.get("dialogue", [])]
    write_table(dialogue, ["时间段", "镜号", "说话人", "台词", "表演/配音语气", "字幕执行"], drows, [18, 9, 18, 58, 34, 34], row_height=54)

    timeline = wb.create_sheet("剪辑时间轴")
    lrows = [[value(x, "time_range"), value(x, "video"), value(x, "dialogue"), value(x, "sound"), value(x, "emotion"), value(x, "transition"), value(x, "completion")] for x in payload.get("timeline", [])]
    write_table(timeline, ["时间段", "画面轨", "对白/字幕轨", "音乐音效轨", "情绪目标", "转场与剪辑", "完成标准"], lrows, [22, 34, 62, 54, 30, 46, 52], row_height=82)
    finish(wb, output, f"{value(payload.get('video', {}), 'title', '视频')}—模仿复刻版")
    return expected_images, embedded


def main():
    parser = argparse.ArgumentParser(description="Build learning-analysis and replication Excel workbooks from analysis JSON.")
    parser.add_argument("analysis_json", type=Path)
    parser.add_argument("--mode", choices=("learning", "replica", "both"), default="both")
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--basename")
    args = parser.parse_args()

    source = args.analysis_json.expanduser().resolve()
    payload = json.loads(source.read_text(encoding="utf-8"))
    out = args.output_dir.expanduser().resolve()
    out.mkdir(parents=True, exist_ok=True)
    base = args.basename or value(payload.get("video", {}), "title", source.stem)
    base = re.sub(r'[<>:"/\\|?*]+', "_", base).strip() or "视频拉片"
    results = []

    if args.mode in ("learning", "both"):
        path = out / f"{base}-学习拆解版.xlsx"
        expected, embedded = create_learning(payload, source.parent, path)
        check = load_workbook(path, read_only=False)
        if check.sheetnames != ["学习拆解说明", "学习拉片表", "结构总结", "技巧清单"] or len(check["学习拉片表"]._images) != embedded:
            raise RuntimeError("Learning workbook validation failed")
        results.append({"mode": "learning", "file": str(path), "shots": len(payload.get("shots", [])), "images": embedded, "expected_images": expected})

    if args.mode in ("replica", "both"):
        path = out / f"{base}-模仿复刻版.xlsx"
        expected, embedded = create_replica(payload, source.parent, path)
        check = load_workbook(path, read_only=False)
        if check.sheetnames != ["复刻执行说明", "分镜脚本", "拍摄清单", "台词文案", "剪辑时间轴"] or len(check["分镜脚本"]._images) != embedded:
            raise RuntimeError("Replica workbook validation failed")
        results.append({"mode": "replica", "file": str(path), "shots": len(payload.get("shots", [])), "images": embedded, "expected_images": expected})

    print(json.dumps(results, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

